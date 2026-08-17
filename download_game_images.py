"""
Game Cover Image Downloader
---------------------------
Put this script in the same folder as games.xlsx and run:

    python download_game_images.py

It reads every sheet in games.xlsx, searches for game cover images using
Bing Image Search through its public HTML results, downloads likely cover art,
and saves images by platform.

Dependencies:
    pip install openpyxl requests beautifulsoup4 pillow

Notes:
- This uses search-engine image results, so no API key is required.
- Search results can occasionally contain an incorrect image. The script
  scores results using the game title + platform and saves the best candidate.
- Respect the image owners' copyrights/licensing when using downloaded images.
"""

from pathlib import Path
import io
import re
import time
import hashlib
from urllib.parse import quote, urlparse

import requests
from bs4 import BeautifulSoup
from PIL import Image
import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE / "games.xlsx"
OUT = BASE / "game-images"
OUT.mkdir(exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0 Safari/537.36"
    )
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

PLATFORM_ALIASES = {
    "PS5": ["PS5", "PlayStation 5"],
    "PS4": ["PS4", "PlayStation 4"],
    "PS3": ["PS3", "PlayStation 3"],
    "XBOX SERIES X": ["Xbox Series X", "Xbox Series"],
    "XBOX ONE": ["Xbox One"],
}

def safe_name(name):
    name = re.sub(r'[<>:"/\\|?*]', "", str(name))
    name = re.sub(r"\s+", " ", name).strip()
    return name[:150]

def image_extension(content_type, url):
    ct = (content_type or "").lower()
    if "png" in ct:
        return ".png"
    if "webp" in ct:
        return ".webp"
    if "jpeg" in ct or "jpg" in ct:
        return ".jpg"
    ext = Path(urlparse(url).path).suffix.lower()
    return ext if ext in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"

def is_probably_image(resp):
    ct = resp.headers.get("Content-Type", "").lower()
    return "image/" in ct

def search_bing_images(game, platform):
    aliases = PLATFORM_ALIASES.get(platform, [platform])
    # Prefer official box/cover terminology.
    query = f'"{game}" "{aliases[0]}" game cover box art'
    url = "https://www.bing.com/images/search?q=" + quote(query)

    try:
        r = SESSION.get(url, timeout=20)
        r.raise_for_status()
    except Exception:
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    results = []

    # Bing commonly stores image metadata in m attributes.
    for a in soup.select("a.iusc"):
        m = a.get("m")
        if not m:
            continue
        try:
            import json
            data = json.loads(m)
        except Exception:
            continue

        img_url = data.get("murl")
        thumb = data.get("turl")
        title = data.get("t")
        if img_url:
            results.append({
                "url": img_url,
                "thumb": thumb,
                "title": title or "",
            })

    return results[:35]

def score_result(game, platform, result):
    text = f"{result.get('title','')} {result.get('url','')}".lower()
    game_words = [w.lower() for w in re.findall(r"[a-zA-Z0-9]+", game) if len(w) > 2]

    score = 0
    for w in game_words:
        if w in text:
            score += 2

    for alias in PLATFORM_ALIASES.get(platform, [platform]):
        if alias.lower() in text:
            score += 5

    # Cover-art related words are useful, but not decisive.
    for w in ("cover", "box", "boxart", "box-art", "front"):
        if w in text:
            score += 2

    # Penalize obvious unrelated formats.
    for w in ("wallpaper", "fanart", "screenshot", "logo", "banner"):
        if w in text:
            score -= 3

    return score

def download_candidate(url):
    try:
        r = SESSION.get(url, timeout=20, stream=True)
        r.raise_for_status()
        if not is_probably_image(r):
            return None

        data = r.content
        if len(data) < 10_000:
            return None

        img = Image.open(io.BytesIO(data))
        img.verify()

        # Re-open after verify.
        img = Image.open(io.BytesIO(data))
        w, h = img.size

        # Reject tiny images and extremely wide banners.
        if w < 300 or h < 300:
            return None

        ratio = w / h
        if ratio > 2.2 or ratio < 0.35:
            return None

        return data, w, h, r.headers.get("Content-Type", "")
    except Exception:
        return None

def save_best(game, platform):
    platform_dir = OUT / safe_name(platform)
    platform_dir.mkdir(parents=True, exist_ok=True)

    base_name = safe_name(game)
    existing = list(platform_dir.glob(base_name + ".*"))
    if existing:
        return str(existing[0]), "EXISTS"

    results = search_bing_images(game, platform)
    results.sort(key=lambda x: score_result(game, platform, x), reverse=True)

    for candidate in results:
        downloaded = download_candidate(candidate["url"])
        if not downloaded:
            continue

        data, w, h, ct = downloaded
        ext = image_extension(ct, candidate["url"])
        path = platform_dir / (base_name + ext)
        path.write_bytes(data)

        return str(path), "DOWNLOADED"

    return "", "FAILED"

def main():
    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found.")
        print("Put games.xlsx in the same folder as this script.")
        return

    wb = openpyxl.load_workbook(XLSX)
    report = []

    for ws in wb.worksheets:
        platform = ws.title.strip()
        print(f"\n===== {platform} =====")

        # Find the first useful text column. Prefer a header named Game/Product/Title.
        headers = {}
        for cell in ws[1]:
            if cell.value is not None:
                headers[str(cell.value).strip().lower()] = cell.column

        game_col = None
        for key in ("game", "games", "title", "product", "product name", "name"):
            if key in headers:
                game_col = headers[key]
                break

        if game_col is None:
            # Fallback: first column.
            game_col = 1

        for row in range(2, ws.max_row + 1):
            game = ws.cell(row=row, column=game_col).value
            if not game:
                continue

            game = str(game).strip()
            print(f"[{platform}] {game}")

            path, status = save_best(game, platform)
            print(f"    -> {status} {path}")

            report.append({
                "Platform": platform,
                "Game": game,
                "Image Path": path,
                "Status": status,
            })

            # Be polite to the search engine.
            time.sleep(1.0)

    # Create a report workbook.
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = "Image Results"

    cols = ["Platform", "Game", "Image Path", "Status"]
    report_ws.append(cols)
    for item in report:
        report_ws.append([item[c] for c in cols])

    report_path = BASE / "image_download_report.xlsx"
    report_wb.save(report_path)

    print("\n====================================")
    print("DONE")
    print(f"Images: {OUT}")
    print(f"Report: {report_path}")
    print("====================================")

if __name__ == "__main__":
    main()
