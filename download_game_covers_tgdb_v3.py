"""
ACCURATE GAME COVER DOWNLOADER — TheGamesDB
============================================
Reads games.xlsx and downloads FRONT BOX ART only.

This version fixes:
- Wrong/obsolete hard-coded platform IDs
- PS5 sheet header being on a different row
- GAME_NOT_FOUND caused by wrong platform IDs
- Random screenshots/wallpapers
- Weak title matching

It dynamically gets the platform ID from TheGamesDB by platform name,
then searches the game and verifies the platform before downloading.

Install:
    python -m pip install openpyxl requests pillow

Run:
    python download_game_covers_tgdb_v3.py
"""

from pathlib import Path
import io
import re
import time
import requests
import openpyxl
from PIL import Image

API_KEY = "abce02c497f99cefe852bf93f38ac57496c703399dd86fdf5807d89cd88b7537"
BASE = "https://api.thegamesdb.net/v1"
CDN = "https://cdn.thegamesdb.net/images/original/"

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "games.xlsx"
OUT = ROOT / "game-covers"
REPORT = ROOT / "image_download_report.xlsx"

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 GameCoverDownloader/3.0"
})

PLATFORM_NAMES = {
    "PS5": "Sony Playstation 5",
    "PS4": "Sony Playstation 4",
    "PS3": "Sony Playstation 3",
    "XBOX SERIES X": "Microsoft Xbox Series X",
    "XBOX ONE": "Microsoft Xbox One",
}


def clean_name(s):
    s = re.sub(r'[<>:"/\\|?*]', "", str(s))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:160]


def norm(s):
    s = str(s).lower()
    s = s.replace("&", "and")
    s = s.replace("’", "'")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def api(endpoint, params):
    p = dict(params)
    p["apikey"] = API_KEY

    try:
        r = SESSION.get(f"{BASE}/{endpoint}", params=p, timeout=30)

        if r.status_code != 200:
            print(f"    API HTTP {r.status_code}: {r.text[:300]}")
            return None

        data = r.json()

        if data.get("status") not in (None, "Success"):
            print(f"    API STATUS: {data.get('status')}")

        if data.get("code") not in (None, 200):
            print(f"    API CODE: {data.get('code')}")

        return data

    except Exception as e:
        print(f"    API ERROR: {e}")
        return None


def get_platform_id(platform_name):
    """
    Don't hard-code platform IDs. TheGamesDB platform IDs can change/expand.
    Ask the API directly by platform name.
    """
    data = api(
        "Platforms/ByPlatformName",
        {"name": platform_name}
    )

    if not data:
        return None

    platforms = data.get("data", {}).get("platforms", [])

    if isinstance(platforms, dict):
        platforms = list(platforms.values())

    wanted = norm(platform_name)

    # Exact match first.
    for p in platforms:
        name = p.get("name", "")
        if norm(name) == wanted:
            return p.get("id")

    # Then contains match.
    for p in platforms:
        name = p.get("name", "")
        if wanted in norm(name) or norm(name) in wanted:
            return p.get("id")

    return None


def get_game_candidates(game_name, platform_id):
    """
    Search with the current v1.1 endpoint.
    Include platform and boxart metadata.
    """
    data = api(
        "Games/ByGameName",
        {
            "name": game_name,
            "filter[platform]": str(platform_id),
            "include": "boxart,platform",
            "fields": "platform,alternates,overview,last_updated",
            "page": 1
        }
    )

    if not data:
        return []

    return data.get("data", {}).get("games", []) or []


def title_score(query, candidate):
    q = norm(query)
    c = norm(candidate)

    if not q or not c:
        return -999

    if q == c:
        return 1000

    qw = set(q.split())
    cw = set(c.split())

    common = len(qw & cw)
    score = common * 30

    if q in c:
        score += 300

    # Avoid accepting something like "Sekiro: Shadows Die Twice Demo"
    extra = len(cw - qw)
    score -= extra * 12

    # Very strong penalty for obvious editions that aren't the requested title.
    bad_terms = {
        "demo", "trailer", "soundtrack", "dlc", "episode",
        "theme", "avatar", "bundle", "collection", "pack",
        "beta", "trial"
    }
    score -= sum(40 for x in c.split() if x in bad_terms)

    return score


def choose_game(game_name, candidates, platform_id):
    valid = []

    for g in candidates:
        gid = g.get("id")
        title = g.get("game_title") or ""

        # Platform can be an integer OR an embedded object depending on API version.
        gp = g.get("platform")
        if isinstance(gp, dict):
            gp_id = gp.get("id")
        else:
            gp_id = gp

        # If platform info exists, enforce it.
        if gp_id is not None:
            try:
                if int(gp_id) != int(platform_id):
                    continue
            except Exception:
                pass

        score = title_score(game_name, title)
        if gid and score >= 25:
            valid.append((score, g))

    if not valid:
        return None

    valid.sort(key=lambda x: x[0], reverse=True)

    # Don't accept a fuzzy result when it is substantially different.
    return valid[0][1]


def get_front_cover(game_id):
    """
    TheGamesDB's Images endpoint returns:
    data.images["GAME_ID"] -> list of image objects
    with type=boxart, side=front, filename=...
    """
    data = api(
        "Games/Images",
        {
            "games_id": str(game_id),
            "filter[type]": "boxart",
            "page": 1
        }
    )

    if not data:
        return None

    images = data.get("data", {}).get("images", {})

    if not isinstance(images, dict):
        return None

    items = images.get(str(game_id), [])

    if isinstance(items, dict):
        items = [items]

    fronts = [
        x for x in items
        if isinstance(x, dict)
        and x.get("type") == "boxart"
        and str(x.get("side", "")).lower() == "front"
        and x.get("filename")
    ]

    if not fronts:
        return None

    # Prefer the highest resolution.
    def resolution_score(x):
        m = re.match(r"(\d+)x(\d+)", str(x.get("resolution", "")))
        if not m:
            return 0
        return int(m.group(1)) * int(m.group(2))

    fronts.sort(key=resolution_score, reverse=True)

    return fronts[0]


def download_cover(filename):
    url = CDN + filename.lstrip("/")

    try:
        r = SESSION.get(url, timeout=30)
        r.raise_for_status()

        img = Image.open(io.BytesIO(r.content))
        img.verify()

        img = Image.open(io.BytesIO(r.content))
        w, h = img.size

        # Strict front-cover sanity checks.
        if w < 300 or h < 400:
            return None

        ratio = w / h

        # A real front cover is portrait. Reject banners/wallpapers.
        if ratio >= 0.95 or ratio < 0.35:
            return None

        return r.content, img.format or "JPEG", w, h

    except Exception:
        return None


def find_game_column(ws):
    # Search first 10 rows, not just row 2.
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value and norm(value) in {"game", "games", "title", "product", "product name"}:
                return row, col

    return None, None


def process(game, platform, platform_id):
    folder = OUT / clean_name(platform)
    folder.mkdir(parents=True, exist_ok=True)

    base = clean_name(game)

    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        existing = folder / (base + ext)
        if existing.exists():
            return str(existing), "EXISTS", ""

    candidates = get_game_candidates(game, platform_id)

    if not candidates:
        return "", "GAME_NOT_FOUND", ""

    selected = choose_game(game, candidates, platform_id)

    if not selected:
        # One retry without platform filter.
        # This catches cases where TGDB has incomplete platform metadata.
        data = api(
            "Games/ByGameName",
            {
                "name": game,
                "include": "boxart,platform",
                "fields": "platform,alternates,overview,last_updated",
                "page": 1
            }
        )

        candidates2 = []
        if data:
            candidates2 = data.get("data", {}).get("games", []) or []

        selected = choose_game(game, candidates2, platform_id)

    if not selected:
        return "", "NO_CONFIDENT_MATCH", ""

    game_id = selected.get("id")
    matched = selected.get("game_title", "")

    print(f"    matched: {matched} | id={game_id}")

    cover = get_front_cover(game_id)

    if not cover:
        return "", "NO_FRONT_COVER", matched

    result = download_cover(cover["filename"])

    if not result:
        return "", "INVALID_FRONT_COVER", matched

    content, fmt, w, h = result

    ext = {
        "JPEG": ".jpg",
        "JPG": ".jpg",
        "PNG": ".png",
        "WEBP": ".webp",
    }.get(str(fmt).upper(), ".jpg")

    path = folder / (base + ext)
    path.write_bytes(content)

    return str(path), "DOWNLOADED", matched


def main():
    if not XLSX.exists():
        print("ERROR: games.xlsx same folder mein nahi hai.")
        input("Press Enter...")
        return

    OUT.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    report_rows = []
    total = 0

    # Resolve platform IDs once.
    platform_ids = {}

    print("Resolving TheGamesDB platform IDs...\n")

    for sheet in wb.worksheets:
        key = sheet.title.strip().upper()

        if key not in PLATFORM_NAMES:
            continue

        platform_name = PLATFORM_NAMES[key]
        pid = get_platform_id(platform_name)

        if pid:
            platform_ids[key] = pid
            print(f"{key}: {platform_name} -> ID {pid}")
        else:
            print(f"{key}: PLATFORM NOT FOUND")

        time.sleep(0.4)

    for ws in wb.worksheets:
        platform = ws.title.strip().upper()

        if platform not in PLATFORM_NAMES:
            continue

        print(f"\n========== {platform} ==========")

        platform_id = platform_ids.get(platform)

        if not platform_id:
            print("Skipping: platform ID unavailable.")
            continue

        header_row, game_col = find_game_column(ws)

        if not game_col:
            print("GAME column nahi mila.")
            continue

        print(f"GAME column found: row={header_row}, column={game_col}")

        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row=row, column=game_col).value

            if value is None or not str(value).strip():
                continue

            game = str(value).strip()
            total += 1

            print(f"[{total}] {game}")

            path, status, matched = process(
                game,
                platform,
                platform_id
            )

            print(f"    {status}")
            if path:
                print(f"    saved: {path}")

            report_rows.append([
                platform,
                game,
                matched,
                path,
                status
            ])

            # Small delay to avoid hitting rate limits.
            time.sleep(0.35)

    # Report
    rw = openpyxl.Workbook()
    rws = rw.active
    rws.title = "Image Results"

    rws.append([
        "Platform",
        "Game",
        "TheGamesDB Match",
        "Image Path",
        "Status"
    ])

    for row in report_rows:
        rws.append(row)

    rw.save(REPORT)

    print("\n==============================================")
    print("DONE")
    print(f"Games processed: {total}")
    print(f"Images: {OUT}")
    print(f"Report: {REPORT}")
    print("==============================================")

    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
