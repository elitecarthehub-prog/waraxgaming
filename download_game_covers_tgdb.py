"""
TheGamesDB Game Cover Downloader
================================
Put this script in the SAME folder as games.xlsx and run it.

Install once:
    python -m pip install openpyxl requests pillow

Then:
    python download_game_covers_tgdb.py

The API key is already configured below.
"""

from pathlib import Path
from urllib.parse import quote
import io
import re
import time
import requests
import openpyxl
from PIL import Image

# =========================
# CONFIG
# =========================
API_KEY = "abce02c497f99cefe852bf93f38ac57496c703399dd86fdf5807d89cd88b7537"

BASE_DIR = Path(__file__).resolve().parent
XLSX = BASE_DIR / "games.xlsx"
OUTPUT = BASE_DIR / "game-covers"

API = "https://api.thegamesdb.net/v1"

HEADERS = {
    "User-Agent": "GameCoverDownloader/2.0"
}

# TheGamesDB platform IDs.
# These are intentionally explicit so a PS5 game doesn't get matched to
# another console just because the title is the same.
PLATFORM_IDS = {
    "PS5": 4959,
    "PS4": 4917,
    "PS3": 4918,
    "XBOX SERIES X": 4947,
    "XBOX ONE": 4920,
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


def clean_filename(text):
    text = re.sub(r'[<>:"/\\|?*]', "", str(text))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:160]


def api_get(endpoint, params):
    params = dict(params)
    params["apikey"] = API_KEY
    url = f"{API}/{endpoint}"

    try:
        r = SESSION.get(url, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        return data
    except Exception as e:
        print(f"    API ERROR: {e}")
        return None


def get_games(game_name, platform_id):
    """
    Search TheGamesDB by exact-ish game name and platform.
    We request a reasonably large result set so the local matcher can
    choose the closest title.
    """
    data = api_get(
        "Games/ByGameName",
        {
            "name": game_name,
            "filter[platform]": platform_id,
            "fields": "players,publishers,genres,overview,last_updated",
            "include": "boxart",
            "page": 1,
            "limit": 20,
        },
    )

    if not data:
        return []

    return data.get("data", {}).get("games", []) or []


def normalize_title(s):
    s = str(s).lower()
    s = s.replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def title_score(query, candidate):
    q = normalize_title(query)
    c = normalize_title(candidate)

    if q == c:
        return 1000

    q_words = set(q.split())
    c_words = set(c.split())

    if not q_words:
        return 0

    common = len(q_words & c_words)
    score = common * 20

    # Strong bonus for exact phrase.
    if q in c:
        score += 250

    # Penalize candidates that have many extra words.
    score -= max(0, len(c_words) - len(q_words)) * 4

    return score


def choose_game(game_name, games):
    if not games:
        return None

    scored = []
    for g in games:
        name = g.get("game_title") or g.get("name") or ""
        score = title_score(game_name, name)
        scored.append((score, g))

    scored.sort(key=lambda x: x[0], reverse=True)

    # Do not blindly accept a poor fuzzy match.
    best_score, best = scored[0]
    if best_score < 20:
        return None

    return best


def get_front_boxart(game_id):
    data = api_get(
        "Games/Images",
        {
            "games_id": game_id,
            "filter[type]": "boxart",
            "filter[side]": "front",
            "resolution": "original",
        },
    )

    if not data:
        return []

    images = data.get("data", {}).get("images", {}) or {}

    # API responses can vary slightly by version; handle common forms.
    if isinstance(images, list):
        return images

    if isinstance(images, dict):
        for key in ("boxart", "games", str(game_id)):
            value = images.get(key)
            if isinstance(value, list):
                return value

        # Some responses contain an object keyed by game id.
        value = images.get(str(game_id))
        if isinstance(value, dict):
            for key in ("boxart", "images"):
                if isinstance(value.get(key), list):
                    return value[key]

    return []


def extract_image_url(image):
    if not isinstance(image, dict):
        return None

    # Possible API fields.
    for key in ("filename", "url", "image", "original"):
        value = image.get(key)
        if isinstance(value, str) and value.strip():
            value = value.strip()

            if value.startswith("http://") or value.startswith("https://"):
                return value

            # TheGamesDB image paths are commonly relative to the image base.
            if value.startswith("/"):
                return "https://cdn.thegamesdb.net" + value

            return "https://cdn.thegamesdb.net/images/original/" + value

    return None


def download_image(url):
    try:
        r = SESSION.get(url, timeout=30)
        r.raise_for_status()

        img = Image.open(io.BytesIO(r.content))
        img.verify()

        img = Image.open(io.BytesIO(r.content))
        w, h = img.size

        # Cover-art sanity check.
        if w < 300 or h < 400:
            return None

        # Game front covers should normally be portrait.
        ratio = w / h
        if ratio > 0.95:
            return None

        return r.content, img.format or "JPEG", w, h

    except Exception:
        return None


def process_game(game_name, platform):
    platform_id = PLATFORM_IDS.get(platform)
    if not platform_id:
        return "", "UNKNOWN_PLATFORM", ""

    platform_dir = OUTPUT / clean_filename(platform)
    platform_dir.mkdir(parents=True, exist_ok=True)

    filename = clean_filename(game_name)

    # Don't download again if already present.
    existing = []
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        p = platform_dir / (filename + ext)
        if p.exists():
            return str(p), "EXISTS", ""

    games = get_games(game_name, platform_id)

    if not games:
        return "", "GAME_NOT_FOUND", ""

    selected = choose_game(game_name, games)
    if not selected:
        return "", "NO_CONFIDENT_MATCH", ""

    game_id = selected.get("id") or selected.get("game_id")
    matched_name = selected.get("game_title") or selected.get("name") or ""

    if not game_id:
        return "", "NO_GAME_ID", matched_name

    print(f"    matched: {matched_name} (ID {game_id})")

    images = get_front_boxart(game_id)
    if not images:
        return "", "NO_FRONT_COVER", matched_name

    # Prefer original/high-resolution image.
    candidates = []
    for image in images:
        url = extract_image_url(image)
        if url:
            candidates.append((image, url))

    # Try candidates until a valid portrait cover is found.
    for image, url in candidates:
        result = download_image(url)
        if not result:
            continue

        content, fmt, w, h = result

        ext = {
            "JPEG": ".jpg",
            "JPG": ".jpg",
            "PNG": ".png",
            "WEBP": ".webp",
        }.get(str(fmt).upper(), ".jpg")

        path = platform_dir / (filename + ext)
        path.write_bytes(content)

        return str(path), "DOWNLOADED", matched_name

    return "", "INVALID_COVER", matched_name


def main():
    if not XLSX.exists():
        print("ERROR: games.xlsx nahi mili.")
        print("games.xlsx ko script ke same folder mein rakho.")
        return

    OUTPUT.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(XLSX)
    report = []

    total = 0
    for ws in wb.worksheets:
        platform = ws.title.strip().upper()

        if platform not in PLATFORM_IDS:
            print(f"Skipping unknown sheet: {ws.title}")
            continue

        # Find GAME column.
        game_col = None
        for cell in ws[2]:
            if str(cell.value).strip().upper() == "GAME":
                game_col = cell.column
                break

        if game_col is None:
            print(f"GAME column nahi mila: {ws.title}")
            continue

        print(f"\n========== {platform} ==========")

        for row in range(3, ws.max_row + 1):
            value = ws.cell(row=row, column=game_col).value

            if not value:
                continue

            game = str(value).strip()
            total += 1

            print(f"[{total}] {game}")

            path, status, matched = process_game(game, platform)

            print(f"    {status}")
            if path:
                print(f"    saved: {path}")

            report.append([
                platform,
                game,
                matched,
                path,
                status,
            ])

            # Avoid hammering the API.
            time.sleep(0.35)

    # Save report.
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = "Image Results"

    report_ws.append([
        "Platform",
        "Game",
        "TheGamesDB Match",
        "Image Path",
        "Status",
    ])

    for row in report:
        report_ws.append(row)

    report_path = BASE_DIR / "image_download_report.xlsx"
    report_wb.save(report_path)

    print("\n========================================")
    print("DONE")
    print(f"Total games processed: {total}")
    print(f"Images folder: {OUTPUT}")
    print(f"Report: {report_path}")
    print("========================================")


if __name__ == "__main__":
    main()
